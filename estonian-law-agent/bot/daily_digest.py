"""
Eesti õiguse agent — igapäevane Riigi Teataja seire (GitHub Actions edition).

Each run: query the Riigi Teataja search API for laws published since the last
run, detect newly effective consolidated versions of the watched laws
(config.yaml), summarize each new act in Estonian with Claude, and send the
digest to Telegram as an .xlsx document (one row per act; separate columns for
the recap and the riigiteataja.ee source link) with a short Estonian caption.

Env vars (GitHub Actions secrets):
  TELEGRAM_BOT_TOKEN, ANTHROPIC_API_KEY
Optional:
  ALLOWED_CHAT_ID — the owner's chat, used as the digest destination

Usage:
  daily_digest.py                 normal run (only acts at 10:00 Europe/Tallinn)
  daily_digest.py --force         bypass the 10:00 local-time guard
  daily_digest.py --dry-run       print the digest, no Telegram, no state write
  daily_digest.py --since DATE    override the "new since" date (YYYY-MM-DD)
  daily_digest.py --seed          write initial state/last_check.json and exit
"""

import argparse
import json
import os
import re
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

API = "https://www.riigiteataja.ee/api/oigusakt_otsing/1/otsi"
AKT_URL = "https://www.riigiteataja.ee/akt/{}"
AKT_XML_URL = "https://www.riigiteataja.ee/public-api/api/v1/akt/{}/blob-xml"
PAGE_SIZE = 500
DEFAULT_MAX_SCAN_PAGES = 40   # safety cap for deep backfills (~20k acts)
XML_TRUNCATE = 30_000   # chars of act text handed to the summary model
MAX_SUMMARY_TOKENS = 400
TALLINN = ZoneInfo("Europe/Tallinn")

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_FILE = BASE_DIR / "config.yaml"
STATE_FILE = BASE_DIR / "state" / "last_check.json"


def parse_globaal_id(gid) -> date | None:
    """Modern globaalID encodes the publication notation: <RT part><DDMM><YYYY><seq>.

    E.g. 116072026004 = RT I, 16.07.2026, 4. Legacy short IDs carry no date.
    """
    s = str(gid)
    if len(s) != 12 or s[0] not in "1234":
        return None
    try:
        return date(int(s[5:9]), int(s[3:5]), int(s[1:3]))
    except ValueError:
        return None


def api_search(**params) -> dict:
    r = requests.get(f"{API}?{urlencode(params)}", timeout=60)
    r.raise_for_status()
    return r.json()


def fetch_new_acts(since: date, seen_ids: set, scope: dict,
                   max_pages: int = DEFAULT_MAX_SCAN_PAGES) -> list[dict]:
    """Return acts published on/after `since`, newest publication date first.

    The API has no publication-date filter and its result ordering is only
    loosely chronological (documents are interleaved by indexing order), so
    a partial scan can miss acts. The laws-only scope is small (~10 pages),
    so scan every page backwards from the end — bounded by `max_pages` as a
    safety cap — and filter by the date embedded in globaalID.
    """
    base = {"limiit": PAGE_SIZE, "tekst": "algtekst", **scope}
    total = api_search(**base, leht=1)["metaandmed"]["kokku"]
    last_page = max(1, -(-total // PAGE_SIZE))
    first_page = max(1, last_page - max_pages + 1)

    acts = []
    for page in range(last_page, first_page - 1, -1):
        for act in api_search(**base, leht=page).get("aktid", []):
            pub = parse_globaal_id(act["globaalID"])
            if pub and pub >= since and act["globaalID"] not in seen_ids:
                act["avaldamise_kp"] = pub.isoformat()
                acts.append(act)
    print(f"Skaneeritud lehti: {last_page - first_page + 1} (kokku {last_page})")
    acts.sort(key=lambda a: (a["avaldamise_kp"], a["globaalID"]), reverse=True)
    return acts


def fetch_current_versions(lyhendid: list[str], today: date) -> dict:
    """For each watched abbreviation return the consolidated text in force today."""
    versions = {}
    for lyhend in lyhendid:
        try:
            aktid = api_search(limiit=5, lyhend=lyhend, kehtiv=today.isoformat()).get("aktid", [])
        except Exception as e:
            print(f"Hoiatus: {lyhend} päring ebaõnnestus: {e}", file=sys.stderr)
            continue
        if aktid:
            act = aktid[0]
            versions[lyhend] = {
                "globaalID": act["globaalID"],
                "pealkiri": act["pealkiri"],
                "kehtivuse_algus": (act.get("kehtivus") or {}).get("algus"),
            }
    return versions


def fetch_act_text(gid) -> str:
    """Plain text of an act, roughly extracted from its XML, for summarization."""
    r = requests.get(AKT_XML_URL.format(gid), timeout=60)
    r.raise_for_status()
    text = re.sub(r"<[^>]+>", " ", r.text)
    return re.sub(r"\s+", " ", text).strip()[:XML_TRUNCATE]


def summarize(act: dict, model: str) -> str | None:
    """2–3 lauseline eestikeelne kokkuvõte; None kui kokkuvõtet ei õnnestu teha."""
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return None
    try:
        from anthropic import Anthropic

        text = fetch_act_text(act["globaalID"])
        resp = Anthropic().messages.create(
            model=model,
            max_tokens=MAX_SUMMARY_TOKENS,
            system=(
                "Oled Eesti õiguse ekspert. Koosta avaldatud õigusaktist 2–3 lauseline "
                "eestikeelne kokkuvõte: mis muutub või kehtestatakse ja kellele see on "
                "oluline. Kasuta AINULT kasutaja sõnumis antud akti teksti (Riigi "
                "Teataja) — ära lisa fakte muudest allikatest ega oma taustateadmistest. "
                "Vasta ainult eesti keeles, ilma sissejuhatuseta."
            ),
            messages=[{
                "role": "user",
                "content": f"Akt: {act['pealkiri']} ({act['valjaandja']})\n\nTekst:\n{text}",
            }],
        )
        return "".join(b.text for b in resp.content if b.type == "text").strip()
    except Exception as e:
        print(f"Hoiatus: kokkuvõte ebaõnnestus ({act['globaalID']}): {e}", file=sys.stderr)
        return None


def watched_title_stems(versions: dict) -> dict:
    """lyhend -> lowercase full title, matched as substring against new act titles.

    Estonian genitive extends the nominative ("võlaõigusseaduse muutmise seadus"
    contains "võlaõigusseadus"), so plain containment catches amendment acts.
    """
    return {ly: v["pealkiri"].lower() for ly, v in versions.items() if v.get("pealkiri")}


def digest_rows(watched_hits, other_acts, new_versions, today) -> list[dict]:
    """Flatten the digest into spreadsheet rows: watched hits first, then the
    rest (both newest first), then newly effective consolidated versions."""
    rows = []
    for act in watched_hits + other_acts:
        rows.append({
            "watched": act in watched_hits,
            "avaldatud": act["avaldamise_kp"],
            "pealkiri": act["pealkiri"],
            "valjaandja": act.get("valjaandja") or "",
            "joustub": (act.get("kehtivus") or {}).get("algus") or "",
            "kokkuvote": act.get("kokkuvote") or "",
            "link": AKT_URL.format(act["globaalID"]),
        })
    for ly, v in new_versions.items():
        rows.append({
            "watched": True,
            "avaldatud": today.isoformat(),
            "pealkiri": f"{v['pealkiri']} ({ly}) — uus terviktekst",
            "valjaandja": "",
            "joustub": v["kehtivuse_algus"] or "",
            "kokkuvote": f"Jälgitava seaduse uus terviktekst kehtib alates "
                         f"{v['kehtivuse_algus']}.",
            "link": AKT_URL.format(v["globaalID"]),
        })
    return rows


def build_workbook(rows: list[dict], today: date) -> Path:
    """Write the digest as an .xlsx file and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Riigi Teataja {today.strftime('%d.%m.%Y')}"

    headers = ["⭐", "Avaldatud", "Pealkiri", "Väljaandja", "Jõustub",
               "Kokkuvõte", "Link"]
    widths = [4, 12, 55, 22, 12, 80, 42]
    ws.append(headers)
    for col, width in enumerate(widths, start=1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    for r, row in enumerate(rows, start=2):
        ws.append(["⭐" if row["watched"] else "", row["avaldatud"],
                   row["pealkiri"], row["valjaandja"], row["joustub"],
                   row["kokkuvote"], row["link"]])
        for col in (3, 6):
            ws.cell(row=r, column=col).alignment = wrap
        link_cell = ws.cell(row=r, column=7)
        link_cell.hyperlink = row["link"]
        link_cell.font = link_font

    path = Path(tempfile.gettempdir()) / f"Riigi_Teataja_{today.isoformat()}.xlsx"
    wb.save(path)
    return path


def build_caption(rows: list[dict], today: date) -> str:
    """Short Estonian caption for the Telegram document (limit 1024 chars)."""
    watched = [r for r in rows if r["watched"]]
    caption = (f"⚖️ Riigi Teataja seire {today.strftime('%d.%m.%Y')} — "
               f"{len(rows)} kirjet, neist ⭐ jälgitavaid: {len(watched)}. "
               f"Kokkuvõtted ja lingid manuses.")
    for r in watched:
        line = f"\n⭐ {r['pealkiri']}"
        if len(caption) + len(line) > 1000:
            caption += "\n⭐ …"
            break
        caption += line
    return caption


def send_telegram_document(path: Path, caption: str):
    tg = f"https://api.telegram.org/bot{os.environ['TELEGRAM_BOT_TOKEN'].strip()}"
    with path.open("rb") as f:
        requests.post(
            f"{tg}/sendDocument",
            data={"chat_id": os.environ["ALLOWED_CHAT_ID"], "caption": caption},
            files={"document": (path.name, f,
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet")},
            timeout=120,
        ).raise_for_status()


def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"last_date": None, "seen_ids": [], "redaktsioonid": {}}


def save_state(state: dict):
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true", help="skip the 10:00 local-time guard")
    ap.add_argument("--dry-run", action="store_true", help="print digest, no Telegram/state")
    ap.add_argument("--since", help="override new-since date (YYYY-MM-DD)")
    ap.add_argument("--seed", action="store_true", help="write initial state and exit")
    args = ap.parse_args()

    now = datetime.now(TALLINN)
    today = now.date()
    force = args.force or os.environ.get("FORCE_RUN") == "true"
    if not (force or args.dry_run or args.seed) and now.hour != 10:
        print(f"Kell on {now:%H:%M} (Europe/Tallinn), mitte 10 — jätan vahele.")
        return

    config = yaml.safe_load(CONFIG_FILE.read_text())
    scope = {k: str(v) for k, v in (config.get("scope") or {}).items()}
    state = load_state()

    versions = fetch_current_versions(config.get("lyhendid", []), today)

    if args.seed:
        state = {
            "last_date": today.isoformat(),
            "seen_ids": [],
            "redaktsioonid": {ly: v["kehtivuse_algus"] for ly, v in versions.items()},
        }
        save_state(state)
        print(f"Algseis salvestatud: {STATE_FILE}")
        return

    since = date.fromisoformat(args.since or state["last_date"] or
                               (today - timedelta(days=3)).isoformat())
    seen_ids = set(state.get("seen_ids", []))

    new_acts = fetch_new_acts(since, seen_ids, scope,
                              config.get("max_scan_pages", DEFAULT_MAX_SCAN_PAGES))

    # Newly effective consolidated versions of watched laws
    old_versions = state.get("redaktsioonid", {})
    new_versions = {
        ly: v for ly, v in versions.items()
        if v["kehtivuse_algus"] and old_versions.get(ly)
        and v["kehtivuse_algus"] > old_versions[ly]
    }

    # New acts that touch a watched law go to the top of the digest
    stems = watched_title_stems(versions)
    watched_hits = [a for a in new_acts
                    if any(s in a["pealkiri"].lower() for s in stems.values())]
    other_acts = [a for a in new_acts if a not in watched_hits]

    print(f"Uusi akte alates {since}: {len(new_acts)} "
          f"(jälgitavaid: {len(watched_hits)}, uusi redaktsioone: {len(new_versions)})")

    if new_acts or new_versions or not config.get("quiet_days", True):
        model = config.get("summary_model", "claude-sonnet-5")
        for act in new_acts:
            act["kokkuvote"] = summarize(act, model) or ""

        rows = digest_rows(watched_hits, other_acts, new_versions, today)
        workbook = build_workbook(rows, today)
        caption = build_caption(rows, today)

        if args.dry_run:
            print(f"\n{caption}\n\nExceli fail: {workbook} ({len(rows)} rida)")
            for row in rows:
                star = "⭐" if row["watched"] else "  "
                print(f"{star} {row['avaldatud']}  {row['pealkiri'][:60]}  {row['link']}")
            return

        send_telegram_document(workbook, caption)
        print(f"Kokkuvõte saadetud Telegrami Exceli failina ({len(rows)} rida).")
    else:
        if args.dry_run:
            print("Uusi akte pole — vaikne päev.")
            return
        print("Uusi akte pole — sõnumit ei saadeta (quiet_days).")

    cutoff = (today - timedelta(days=14)).isoformat()
    state["last_date"] = today.isoformat()
    state["seen_ids"] = sorted(
        {i for i in seen_ids if (parse_globaal_id(i) or date.min).isoformat() >= cutoff}
        | {a["globaalID"] for a in new_acts}
    )
    state["redaktsioonid"] = {
        ly: v["kehtivuse_algus"] for ly, v in versions.items() if v["kehtivuse_algus"]
    }
    save_state(state)
    print(f"Seis uuendatud: {STATE_FILE}")


if __name__ == "__main__":
    main()
