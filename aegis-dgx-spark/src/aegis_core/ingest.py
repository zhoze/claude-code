#!/usr/bin/env python3
"""Document ingestion pipeline. See docs/part-14-document-pipeline.md.

Runs as aegis-ht, with NO internet access. That is the point: a document
processor that can fetch is an exfiltration channel wearing a uniform. The
only network it touches is loopback — the embedding service (:8001) and the
vector db (:6333), both of which nftables allows for aegis-ht — via plain
REST with stdlib urllib. No qdrant-client, no grpc, no compiled surprises.

Stages: validate -> integrity -> extract -> OCR? -> metadata -> classify
        -> chunk -> embed -> index -> verify

Classification fails closed: an unrecognised document becomes `confidential`
and high-zone. Over-restricting costs an inconvenience; under-restricting
costs the thing this system protects.

Failure semantics: a CONTENT problem (unsupported format, no text layer,
broken file) quarantines the document. An INFRASTRUCTURE problem (embedding
service down, vector db down) fails the run loudly and leaves the document
where it is — quarantining your files because a service was restarting
would be wrong twice.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import shutil
import sys
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path

log = logging.getLogger("aegis.ingest")

KNOWLEDGE = Path("/aegis/knowledge")
QUARANTINE = KNOWLEDGE / ".quarantine"
MAX_BYTES = 200 * 1024 * 1024

TEXT_SUFFIXES = {".txt", ".md", ".csv"}
DOC_SUFFIXES = {".pdf", ".docx", ".xlsx"}
SIDECAR_SUFFIXES = (".meta.json", ".reason")

EMBED_URL = os.environ.get("AEGIS_EMBED_URL",
                           "http://127.0.0.1:8001/v1/embeddings")
QDRANT_URL = os.environ.get("AEGIS_QDRANT_URL", "http://127.0.0.1:6333")
COLLECTION = os.environ.get("AEGIS_INGEST_COLLECTION", "aegis_knowledge")
EMBED_BATCH = 16

# Loopback only; make sure a stray HTTPS_PROXY can never detour us.
_OPENER = urllib.request.build_opener(urllib.request.ProxyHandler({}))


class InfraError(RuntimeError):
    """Embedding service or vector db unavailable — NOT a document problem."""


@dataclass
class DocMeta:
    path: str
    sha256: str
    bytes: int
    imported: str
    classification: str = "confidential"   # fail closed
    trust_zone: str = "high"               # fail closed
    language: str = "unknown"
    title: str = ""
    chunks: int = 0
    collection: str = ""


def quarantine(path: Path, reason: str) -> None:
    QUARANTINE.mkdir(parents=True, exist_ok=True)
    dest = QUARANTINE / path.name
    shutil.move(str(path), dest)
    dest.with_suffix(dest.suffix + ".reason").write_text(
        f"{datetime.now(timezone.utc).isoformat()}\n{reason}\n"
    )
    log.warning("quarantined %s: %s", path.name, reason)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def validate(path: Path) -> str | None:
    if not path.is_file():
        return "not a regular file"
    size = path.stat().st_size
    if size == 0:
        return "empty file"
    if size > MAX_BYTES:
        return f"exceeds size limit ({size} bytes)"
    if path.suffix.lower() not in TEXT_SUFFIXES | DOC_SUFFIXES:
        return f"unsupported format {path.suffix!r}"
    return None


def extract(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in TEXT_SUFFIXES:
        return path.read_text(errors="replace")
    if suffix == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
        if not text.strip():
            # No text layer. OCR is local-only; there is no cloud fallback
            # here by design.
            raise ValueError("no text layer — needs local OCR (not yet wired)")
        return text
    if suffix == ".docx":
        import docx
        return "\n".join(p.text for p in docx.Document(str(path)).paragraphs)
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            rows.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join("" if c is None else str(c) for c in row))
        return "\n".join(rows)
    raise ValueError(f"no extractor for {suffix}")


def chunk(text: str, size: int = 800, overlap: int = 100) -> list[str]:
    """Structure-aware where it can be, fixed-window where it cannot.

    Paragraph boundaries first, because splitting a contract clause in half
    is how retrieval starts returning confident nonsense. A paragraph longer
    than the window (OCR'd tables, minified exports) is split by a sliding
    fixed window — the old version silently emitted it whole, which blew
    past the embedding model's context.

    `size`/`overlap` are in tokens, approximated at 4 chars/token. Chunks
    may exceed the limit by up to one overlap's worth where context is
    carried across a boundary.
    """
    limit = size * 4
    ov = overlap * 4
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]

    pieces: list[str] = []
    for para in paragraphs:
        if len(para) <= limit:
            pieces.append(para)
        else:
            step = max(limit - ov, 1)
            pieces.extend(para[i:i + limit] for i in range(0, len(para), step))

    chunks: list[str] = []
    current = ""
    for piece in pieces:
        if current and len(current) + 2 + len(piece) > limit:
            chunks.append(current)
            current = (current[-ov:] + "\n\n" + piece) if ov else piece
        else:
            current = f"{current}\n\n{piece}" if current else piece
    if current:
        chunks.append(current)
    return chunks


# ---------------------------------------------------------------------------
# Embedding + index (loopback REST, stdlib only)
# ---------------------------------------------------------------------------


def _http(method: str, url: str, payload: dict | None = None,
          timeout: float = 120.0) -> dict:
    data = json.dumps(payload).encode() if payload is not None else None
    req = urllib.request.Request(url, data=data, method=method,
                                 headers={"Content-Type": "application/json"})
    try:
        with _OPENER.open(req, timeout=timeout) as resp:
            body = resp.read()
            return json.loads(body) if body else {}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode(errors="replace")[:300]
        raise InfraError(f"{method} {url} -> {exc.code}: {detail}") from exc
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise InfraError(f"{method} {url} failed: {exc}") from exc


def embed(texts: list[str]) -> list[list[float]]:
    """Batch-embed via the local llama.cpp embedding service."""
    vectors: list[list[float]] = []
    for i in range(0, len(texts), EMBED_BATCH):
        batch = texts[i:i + EMBED_BATCH]
        out = _http("POST", EMBED_URL,
                    {"model": "aegis-embed", "input": batch})
        rows = sorted(out.get("data", []), key=lambda d: d.get("index", 0))
        if len(rows) != len(batch):
            raise InfraError(
                f"embedding service returned {len(rows)} vectors "
                f"for {len(batch)} inputs")
        vectors.extend(r["embedding"] for r in rows)
    return vectors


def ensure_collection(dim: int) -> None:
    url = f"{QDRANT_URL}/collections/{COLLECTION}"
    try:
        info = _http("GET", url)
        have = (info.get("result", {}).get("config", {})
                .get("params", {}).get("vectors", {}).get("size"))
        if have not in (None, dim):
            raise InfraError(
                f"collection {COLLECTION} has dim {have}, embeddings have "
                f"{dim} — you changed the embedding model; re-ingest into a "
                "new collection (the vector db is never the source of truth)")
        return
    except InfraError as exc:
        if "-> 404" not in str(exc):
            raise
    _http("PUT", url, {"vectors": {"size": dim, "distance": "Cosine"}})
    log.info("created collection %s (dim=%d)", COLLECTION, dim)


def upsert(meta: DocMeta, chunks: list[str],
           vectors: list[list[float]]) -> None:
    """Idempotent by construction: point ids derive from (sha256, index), so
    re-ingesting the same file overwrites its own points. The permission
    payload rides with every point — part-13 requires the filter to run
    BEFORE ranking, which Qdrant does with payload filters."""
    points = []
    for i, (text, vec) in enumerate(zip(chunks, vectors)):
        pid = str(uuid.uuid5(uuid.NAMESPACE_URL, f"aegis:{meta.sha256}:{i}"))
        points.append({
            "id": pid,
            "vector": vec,
            "payload": {
                "sha256": meta.sha256,
                "path": meta.path,
                "title": meta.title,
                "chunk_index": i,
                "classification": meta.classification,
                "trust_zone": meta.trust_zone,
                "imported": meta.imported,
                "text": text,
            },
        })
    _http("PUT", f"{QDRANT_URL}/collections/{COLLECTION}/points?wait=true",
          {"points": points})


def ingest_one(path: Path, *, dry_run: bool = False) -> DocMeta | None:
    problem = validate(path)
    if problem:
        if not dry_run:
            quarantine(path, problem)
        return None

    try:
        text = extract(path)
    except InfraError:
        raise
    except Exception as exc:
        if not dry_run:
            quarantine(path, f"extraction failed: {exc}")
        return None

    pieces = chunk(text)
    meta = DocMeta(
        path=str(path),
        sha256=sha256(path),
        bytes=path.stat().st_size,
        imported=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        title=path.stem,
        chunks=len(pieces),
        collection=COLLECTION,
    )

    log.info("%s: %d chars, %d chunks, class=%s zone=%s",
             path.name, len(text), len(pieces), meta.classification,
             meta.trust_zone)

    if dry_run:
        return meta

    vectors = embed(pieces)
    ensure_collection(len(vectors[0]))
    upsert(meta, pieces, vectors)

    sidecar = path.with_suffix(path.suffix + ".meta.json")
    sidecar.write_text(json.dumps(asdict(meta), indent=2))
    return meta


def is_sidecar(path: Path) -> bool:
    return any(path.name.endswith(s) for s in SIDECAR_SUFFIXES)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", type=Path)
    ap.add_argument("--reindex-all", action="store_true",
                    help="re-embed and re-upsert every document under "
                         "/aegis/knowledge (idempotent point ids; quarantine "
                         "and sidecar files are skipped)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)-7s %(message)s")

    if args.reindex_all:
        targets = [p for p in KNOWLEDGE.rglob("*")
                   if p.is_file() and ".quarantine" not in p.parts
                   and not is_sidecar(p)]
    elif args.path:
        targets = [args.path]
    else:
        ap.error("give --path or --reindex-all")

    ok = 0
    try:
        for t in targets:
            if ingest_one(t, dry_run=args.dry_run):
                ok += 1
    except InfraError as exc:
        log.error("infrastructure failure — stopping, documents untouched: %s", exc)
        log.error("check: systemctl status aegis-embed aegis-vectordb")
        return 2

    print(f"\n{ok}/{len(targets)} ingested"
          f"{' (dry run)' if args.dry_run else ''}")
    return 0 if ok == len(targets) else 1


if __name__ == "__main__":
    sys.exit(main())
